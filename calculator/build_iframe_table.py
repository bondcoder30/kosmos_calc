"""
Берёт тортоначинки.xlsx, добавляет 2 колонки:
  - URL  (прямая ссылка на калькулятор торта)
  - IFRAME (готовый сниппет для вставки в редимаг / передачи верстальщику)

Сохраняет рядом как тортоначинки_with_iframes.xlsx.
"""

import sys, io, re, shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).parent
# исходный xlsx лежит в C:\Users\A\Desktop\kosmos\, а текущий __file__ —
# в C:\Users\A\Desktop\kosmos\kosmos_calc\calculator\build_iframe_table.py
# поэтому поднимаемся на 2 уровня (kosmos_calc → kosmos)
SRC  = ROOT.parent.parent / "tortonachinki.xlsx"
DST  = ROOT.parent.parent / "тортоначинки_with_iframes.xlsx"

BASE_URL   = "https://ab-aacoop.github.io/kosmos_calc/calculator/cakes"
BASE_URL_M = "https://ab-aacoop.github.io/kosmos_calc/calculator/cakes-mobile"

# Маппинг: имя в xlsx → (тип папки, id файла) ─────────────────────
CAKE_MAP = {
    "богемская рапсодия":      ("weight", "bohemian-rhapsody"),
    "fairy cake":              ("tiered", "fairy-cake"),
    "big cherry fairy cake":   ("tiered", "big-cherry-fairy-cake"),
    "вавилон":                 ("weight", "babylon"),
    "лебединое озеро":         ("weight", "swan-lake"),
    "фаберже":                 ("fixed",  "faberge"),
    "белль":                   ("tiered", "bell"),
    "люмьер":                  ("tiered", "lumiere"),
    "green day":               ("weight", "green-day"),
    "вишневый сад":            ("weight", "cherry-orchard"),
    "куинджи":                 ("tiered", "kuinji"),
    "дарси":                   ("tiered", "darcy"),
    "shine bright":            ("tiered", "shine-bright"),
    "любовное настроение":     ("weight", "love-in-mood"),
    "тирамису":                ("weight", "tiramisu"),
    "secret garden":           ("fixed",  "secret-garden"),
    "анна":                    ("weight", "anna"),
    "шапито":                  ("tiered", "chapito"),
    "келли":                   ("tiered", "kelly"),
    "ягодные поля навсегда":   ("weight", "berry-fields"),
    "вам письмо":              ("weight", "letter"),
    "орфей":                   ("weight", "orpheus"),
    "blueberry hill":          ("fixed",  "blueberry-hill"),
    "аполлон":                 ("weight", "apollo"),
    "сэйлор мун":              ("weight", "sailor-moon"),
    "sailor moon":             ("weight", "sailor-moon"),
    "dancing queen":           ("weight", "dancing-queen"),
    "тоторо":                  ("weight", "totoro"),
    "элизабет":                ("tiered", "elizabeth"),
    # дополнительно из docx (нет в xlsx, но генерируются)
    "la la land":              ("fixed",  "la-la-land"),
    "fuji":                    ("weight", "fuji"),
}

def normalize(name):
    if not name: return ""
    return re.sub(r"\s+", " ", str(name).strip().lower())

def lookup(name):
    n = normalize(name)
    for key, val in CAKE_MAP.items():
        if normalize(key) == n:
            return val
    return None

def iframe_for(type_, cid):
    url = f"{BASE_URL}/{type_}/{cid}.html"
    return (
        f'<iframe src="{url}" '
        f'style="width:100%;height:100%;border:0;display:block;background:#cfcfcf" '
        f'loading="lazy"></iframe>'
    )

def iframe_for_mobile(type_, cid):
    url = f"{BASE_URL_M}/{type_}/{cid}.html"
    return (
        f'<iframe src="{url}" '
        f'style="width:100%;height:620px;border:0;display:block;background:#cfcfcf" '
        f'loading="lazy"></iframe>'
    )

# ───────────────────────────────────────────────────────────────
shutil.copyfile(SRC, DST)
wb = load_workbook(DST)
ws = wb.active

# Шапка таблицы — первая строка (там колонки начинок). Добавляем 4 справа.
last_col = ws.max_column

URL_COL      = last_col + 1
IFRAME_COL   = last_col + 2
URL_M_COL    = last_col + 3
IFRAME_M_COL = last_col + 4

ws.cell(row=1, column=URL_COL,      value="URL десктоп")
ws.cell(row=1, column=IFRAME_COL,   value="IFRAME десктоп")
ws.cell(row=1, column=URL_M_COL,    value="URL мобильный")
ws.cell(row=1, column=IFRAME_M_COL, value="IFRAME мобильный")

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="D2363C")
for col in (URL_COL, IFRAME_COL, URL_M_COL, IFRAME_M_COL):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

found, missing = [], []
for row_idx in range(2, ws.max_row + 1):
    name = ws.cell(row=row_idx, column=1).value
    if not name or not str(name).strip():
        continue
    match = lookup(name)
    if match:
        type_, cid = match
        url   = f"{BASE_URL}/{type_}/{cid}.html"
        url_m = f"{BASE_URL_M}/{type_}/{cid}.html"
        ws.cell(row=row_idx, column=URL_COL,      value=url)
        ws.cell(row=row_idx, column=IFRAME_COL,   value=iframe_for(type_, cid))
        ws.cell(row=row_idx, column=URL_M_COL,    value=url_m)
        ws.cell(row=row_idx, column=IFRAME_M_COL, value=iframe_for_mobile(type_, cid))
        for col in (URL_COL, IFRAME_COL, URL_M_COL, IFRAME_M_COL):
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="center")
        found.append(str(name).strip())
    else:
        ws.cell(row=row_idx, column=IFRAME_COL,   value="— (нет данных по торту)")
        ws.cell(row=row_idx, column=IFRAME_M_COL, value="— (нет данных по торту)")
        missing.append(str(name).strip())

# Ширина новых колонок
from openpyxl.utils import get_column_letter
ws.column_dimensions[get_column_letter(URL_COL)].width      = 55
ws.column_dimensions[get_column_letter(IFRAME_COL)].width   = 80
ws.column_dimensions[get_column_letter(URL_M_COL)].width    = 55
ws.column_dimensions[get_column_letter(IFRAME_M_COL)].width = 80

wb.save(DST)
print(f"Сохранено: {DST}")
print(f"\nСопоставлено: {len(found)} тортов")
for n in found:
    print(f"  ✓ {n}")
if missing:
    print(f"\nНе найдено в калькуляторе: {len(missing)}")
    for n in missing:
        print(f"  ✗ {n}")
